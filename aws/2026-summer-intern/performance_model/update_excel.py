#!/usr/bin/env python3
"""
Update Results.xlsx with P5en AllReduce roofline model data from nccl_test outputs.
Replicates formulas from the p5en-allreduce-0x7 and p5en-allreduce-0x0 notebooks.
"""
import numpy as np
import re
import math
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ============================================================
# Constants (from notebooks)
# ============================================================
NIC_BW = 50 * 1e3 * 0.975  # 48750 bytes/us
a = np.mean([24.966, 24.815, 24.979])  # 24.92 us (mean RTT for msg <= 32B)
n_gpus_per_node = 8

# Intra-node latency parameters (from 0x0 notebook cell 2)
base_latency_ring_simple = 8.4
base_latency_tree_simple = 8.4
hw_latency_ring_simple = 3.4
hw_latency_tree_simple = 4.0

intra_node_ring_latency = base_latency_ring_simple + (n_gpus_per_node - 1) * hw_latency_ring_simple  # 32.2
intra_node_tree_latency = base_latency_tree_simple + (n_gpus_per_node - 1) * hw_latency_tree_simple  # 36.4
intra_node_ring_latency_ll_ll128 = base_latency_ring_simple  # 8.4
intra_node_tree_latency_ll_ll128 = base_latency_tree_simple  # 8.4

# Protocol chunk sizes
CHUNK_SIZE_LL = 32 * 1024  # 32 KB
CHUNK_SIZE_LL128 = 576000  # bytes

# Single EFA threshold (0x7 only)
SINGLE_EFA_THRESHOLD = 128 * 1024  # 128 KB
NIC_BW_SINGLE_EFA = 50 / 2 * 1e3 * 0.975  # 24375 bytes/us
NIC_BW_DUAL_EFA = NIC_BW  # 48750 bytes/us

# 0x0 Ring Simple threshold
THRESHOLD_LATENCY = 700  # us

# ============================================================
# Data parsing
# ============================================================
def parse_section_data(section_text):
    """Extract (size_bytes, time_us, algbw_GBs, busbw_GBs) from nccl_test section."""
    sizes, times, algbws, busbws = [], [], [], []
    for line in section_text.split('\n'):
        m = re.match(r'\s+(\d+)\s+\d+\s+\w+\s+\w+\s+\S+\s+'
                     r'[\d.]+\s+[\d.]+\s+[\d.]+\s+\d+\s+'
                     r'([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+\d+', line)
        if m:
            sizes.append(int(m.group(1)))
            times.append(float(m.group(2)))
            algbws.append(float(m.group(3)))
            busbws.append(float(m.group(4)))
    return np.array(sizes), np.array(times), np.array(algbws), np.array(busbws)


def parse_file(filepath):
    """Parse nccl_test output file into {(algo, proto, nch): (sizes, times, algbws, busbws)}."""
    with open(filepath, 'r') as f:
        content = f.read()
    sections = re.split(r'={10,}\nConfig:', content)
    results = {}
    for sec in sections:
        cfg_match = re.search(
            r'NCCL_ALGO=(\w+), NCCL_PROTO=(\w+), NCCL_MIN_NCHANNELS=(\d+)', sec)
        if not cfg_match:
            continue
        algo = cfg_match.group(1)
        proto = cfg_match.group(2)
        nch = int(cfg_match.group(3))
        sizes, times, algbws, busbws = parse_section_data(sec)
        if len(sizes) > 0:
            results[(algo, proto, nch)] = (sizes, times, algbws, busbws)
    return results


# ============================================================
# Roofline formulas for 0x7
# ============================================================
def roofline_0x7_ring_simple(sizes, n_nodes, nch):
    """Ring Simple roofline for 0x7 split mask."""
    busbw_factor = 2 * (n_nodes - 1) / n_nodes
    fixed_per_step = (2 * n_nodes - 2) * (a / 2)
    G = 2 * (n_nodes - 1) / (n_nodes * NIC_BW)

    msg_per_gpu_ch = sizes / (n_nodes * nch)
    chunk_size = np.minimum(4 * 1024 * 1024, msg_per_gpu_ch)
    n_chunks = np.ceil(msg_per_gpu_ch / chunk_size).astype(int)
    fixed = fixed_per_step * n_chunks
    T = fixed + sizes * G
    algbw = sizes / (T * 1e-6) / 1e9
    busbw = algbw * busbw_factor
    return T, busbw, busbw_factor


def roofline_0x7_ring_ll(sizes, n_nodes, nch):
    """Ring LL roofline for 0x7 split mask."""
    busbw_factor = 2 * (n_nodes - 1) / n_nodes
    n_pipeline_steps = np.ceil(sizes / (nch * CHUNK_SIZE_LL * n_nodes)).astype(int)
    fixed = 2 * (n_nodes - 1) * (a / 2) * n_pipeline_steps

    # NIC_BW: single EFA if nch==1 and chunk_per_gpu <= 128KB
    chunk_per_gpu = sizes / (nch * n_nodes)
    nic_bw = np.where((nch == 1) & (chunk_per_gpu <= SINGLE_EFA_THRESHOLD),
                      NIC_BW_SINGLE_EFA, NIC_BW_DUAL_EFA)
    G = 2 * 2 * (n_nodes - 1) / (n_nodes * nic_bw)
    T = fixed + sizes * G
    algbw = sizes / (T * 1e-6) / 1e9
    busbw = algbw * busbw_factor
    return T, busbw, busbw_factor


def roofline_0x7_ring_ll128(sizes, n_nodes, nch):
    """Ring LL128 roofline for 0x7 split mask."""
    busbw_factor = 2 * (n_nodes - 1) / n_nodes
    n_pipeline_steps = np.ceil(sizes / (nch * CHUNK_SIZE_LL128 * n_nodes)).astype(int)
    fixed = 2 * (n_nodes - 1) * (a / 2) * n_pipeline_steps

    chunk_per_gpu = sizes / (nch * n_nodes)
    nic_bw = np.where((nch == 1) & (chunk_per_gpu <= SINGLE_EFA_THRESHOLD),
                      NIC_BW_SINGLE_EFA, NIC_BW_DUAL_EFA)
    G = (16.0 / 15.0) * 2 * (n_nodes - 1) / (n_nodes * nic_bw)
    T = fixed + sizes * G
    algbw = sizes / (T * 1e-6) / 1e9
    busbw = algbw * busbw_factor
    return T, busbw, busbw_factor


def roofline_0x7_tree_simple(sizes, n_nodes, nch):
    """Tree Simple roofline for 0x7 split mask."""
    busbw_factor = 2 * (n_nodes - 1) / n_nodes
    fixed = (2 * np.log2(n_nodes) + 1) * a / 2

    # NIC_BW_REAL depends on nch: /3 for 1-2ch, /2 for 4+ch
    if nch <= 2:
        nic_bw_real = NIC_BW / 3
    else:
        nic_bw_real = NIC_BW / 2
    G = 1 / nic_bw_real
    T = fixed + sizes * G
    algbw = sizes / (T * 1e-6) / 1e9
    busbw = algbw * busbw_factor
    return T, busbw, busbw_factor


def roofline_0x7_tree_ll(sizes, n_nodes, nch):
    """Tree LL roofline for 0x7 split mask."""
    busbw_factor = 2 * (n_nodes - 1) / n_nodes
    fixed = (2 * np.log2(n_nodes) + 1) * a / 2

    # NIC_BW: single/dual EFA
    chunk_per_gpu = sizes / (nch * n_nodes)
    nic_bw_base = np.where((nch == 1) & (chunk_per_gpu <= SINGLE_EFA_THRESHOLD),
                           NIC_BW_SINGLE_EFA, NIC_BW_DUAL_EFA)
    # LL: effective BW = NIC_BW / 2 (data + flag)
    nic_bw_eff = nic_bw_base / 2
    # Tree bottleneck: /3 for nch<=2, /2 for nch>=4
    tree_divisor = 3 if nch <= 2 else 2
    bw_real = nic_bw_eff / tree_divisor
    G = 1 / bw_real
    T = fixed + sizes * G
    algbw = sizes / (T * 1e-6) / 1e9
    busbw = algbw * busbw_factor
    return T, busbw, busbw_factor


def roofline_0x7_tree_ll128(sizes, n_nodes, nch):
    """Tree LL128 roofline for 0x7 split mask."""
    busbw_factor = 2 * (n_nodes - 1) / n_nodes
    fixed = (2 * np.log2(n_nodes) + 1) * a / 2

    chunk_per_gpu = sizes / (nch * n_nodes)
    nic_bw_base = np.where((nch == 1) & (chunk_per_gpu <= SINGLE_EFA_THRESHOLD),
                           NIC_BW_SINGLE_EFA, NIC_BW_DUAL_EFA)
    # LL128: effective BW = NIC_BW * 15/16
    nic_bw_eff = nic_bw_base * 15 / 16
    tree_divisor = 3 if nch <= 2 else 2
    bw_real = nic_bw_eff / tree_divisor
    G = 1 / bw_real
    T = fixed + sizes * G
    algbw = sizes / (T * 1e-6) / 1e9
    busbw = algbw * busbw_factor
    return T, busbw, busbw_factor


# ============================================================
# Roofline formulas for 0x0
# ============================================================
def roofline_0x0_ring_simple(sizes, N_nodes, nch_raw):
    """Ring Simple roofline for 0x0 split mask."""
    n_total = N_nodes * 8
    busbw_factor = 2 * (n_total - 1) / n_total
    nch = min(nch_raw, 8) if nch_raw >= 8 else nch_raw
    G_agg = nch * NIC_BW
    G_raw = nch_raw * NIC_BW
    G_per_byte = busbw_factor / G_agg

    fixed_small = (2 * n_total - 1) * a / 2
    fixed_large = (2 * N_nodes - 1) * a / 2

    transfer = sizes / G_raw  # time to transfer through raw channels
    fixed = np.where(transfer < THRESHOLD_LATENCY, fixed_small, fixed_large)
    T = fixed + sizes * G_per_byte
    algbw = sizes / (T * 1e-6) / 1e9
    busbw = algbw * busbw_factor
    return T, busbw, busbw_factor


def roofline_0x0_ring_ll(sizes, N_nodes, nch_raw):
    """Ring LL roofline for 0x0 split mask."""
    n_total = N_nodes * 8
    busbw_factor = 2 * (n_total - 1) / n_total
    nch = min(nch_raw, 8) if nch_raw >= 8 else nch_raw
    G_agg = nch * NIC_BW

    fixed_per_step = (2 * N_nodes - 1) * (a / 2 + intra_node_ring_latency_ll_ll128)
    denom = nch_raw * CHUNK_SIZE_LL * n_total
    n_steps = np.ceil(sizes / denom).astype(int)
    fixed = fixed_per_step * n_steps

    G = 2 * 2 * (n_total - 1) / (n_total * G_agg)
    T = fixed + sizes * G
    algbw = sizes / (T * 1e-6) / 1e9
    busbw = algbw * busbw_factor
    return T, busbw, busbw_factor


def roofline_0x0_ring_ll128(sizes, N_nodes, nch_raw):
    """Ring LL128 roofline for 0x0 split mask."""
    n_total = N_nodes * 8
    busbw_factor = 2 * (n_total - 1) / n_total
    nch = min(nch_raw, 8) if nch_raw >= 8 else nch_raw
    G_agg = nch * NIC_BW

    fixed_per_step = (2 * N_nodes - 1) * (a / 2 + intra_node_ring_latency_ll_ll128)
    denom = nch_raw * CHUNK_SIZE_LL128 * n_total
    n_steps = np.ceil(sizes / denom).astype(int)
    fixed = fixed_per_step * n_steps

    G = (16.0 / 15.0) * 2 * (n_total - 1) / (n_total * G_agg)
    T = fixed + sizes * G
    algbw = sizes / (T * 1e-6) / 1e9
    busbw = algbw * busbw_factor
    return T, busbw, busbw_factor


def roofline_0x0_tree_simple(sizes, N_nodes, nch_raw):
    """Tree Simple roofline for 0x0 split mask."""
    n_total = N_nodes * 8
    busbw_factor = 2 * (n_total - 1) / n_total
    nch_eff = min(nch_raw, 8)

    fixed = (2 * np.log2(N_nodes) + 1) * a / 2 + 2 * intra_node_tree_latency
    nic_bw_real = nch_eff * NIC_BW / 3
    G = 1 / nic_bw_real
    T = fixed + sizes * G
    algbw = sizes / (T * 1e-6) / 1e9
    busbw = algbw * busbw_factor
    return T, busbw, busbw_factor


def roofline_0x0_tree_ll(sizes, N_nodes, nch_raw):
    """Tree LL roofline for 0x0 split mask."""
    # For 4-node case, notebook uses busbw_factor = 2*(N-1)/N (likely a bug),
    # but for 8/16 node case it uses 2*(N*8-1)/(N*8).
    # We replicate the 8/16 node formula for all scales (the correct formula).
    n_total = N_nodes * 8
    busbw_factor = 2 * (n_total - 1) / n_total
    nch_eff = min(nch_raw, 8)

    fixed = (2 * np.log2(N_nodes) + 1) * a / 2 + 2 * intra_node_tree_latency_ll_ll128
    nic_bw_base = NIC_BW / 2  # LL: data + flag
    bw_real = nch_eff * nic_bw_base / 3
    G = 1 / bw_real
    T = fixed + sizes * G
    algbw = sizes / (T * 1e-6) / 1e9
    busbw = algbw * busbw_factor
    return T, busbw, busbw_factor


def roofline_0x0_tree_ll128(sizes, N_nodes, nch_raw):
    """Tree LL128 roofline for 0x0 split mask."""
    n_total = N_nodes * 8
    busbw_factor = 2 * (n_total - 1) / n_total
    nch_eff = min(nch_raw, 8)

    fixed = (2 * np.log2(N_nodes) + 1) * a / 2 + 2 * intra_node_tree_latency_ll_ll128
    nic_bw_base = NIC_BW * 15 / 16  # LL128: 15/16 efficiency
    bw_real = nch_eff * nic_bw_base / 3
    G = 1 / bw_real
    T = fixed + sizes * G
    algbw = sizes / (T * 1e-6) / 1e9
    busbw = algbw * busbw_factor
    return T, busbw, busbw_factor


# ============================================================
# Dispatch roofline computation
# ============================================================
def compute_roofline(split_mask, algo, proto, sizes, n_nodes, nch):
    """Dispatch to appropriate roofline function."""
    if split_mask == '0x7':
        if algo == 'Ring' and proto == 'Simple':
            return roofline_0x7_ring_simple(sizes, n_nodes, nch)
        elif algo == 'Ring' and proto == 'LL':
            return roofline_0x7_ring_ll(sizes, n_nodes, nch)
        elif algo == 'Ring' and proto == 'LL128':
            return roofline_0x7_ring_ll128(sizes, n_nodes, nch)
        elif algo == 'Tree' and proto == 'Simple':
            return roofline_0x7_tree_simple(sizes, n_nodes, nch)
        elif algo == 'Tree' and proto == 'LL':
            return roofline_0x7_tree_ll(sizes, n_nodes, nch)
        elif algo == 'Tree' and proto == 'LL128':
            return roofline_0x7_tree_ll128(sizes, n_nodes, nch)
    elif split_mask == '0x0':
        if algo == 'Ring' and proto == 'Simple':
            return roofline_0x0_ring_simple(sizes, n_nodes, nch)
        elif algo == 'Ring' and proto == 'LL':
            return roofline_0x0_ring_ll(sizes, n_nodes, nch)
        elif algo == 'Ring' and proto == 'LL128':
            return roofline_0x0_ring_ll128(sizes, n_nodes, nch)
        elif algo == 'Tree' and proto == 'Simple':
            return roofline_0x0_tree_simple(sizes, n_nodes, nch)
        elif algo == 'Tree' and proto == 'LL':
            return roofline_0x0_tree_ll(sizes, n_nodes, nch)
        elif algo == 'Tree' and proto == 'LL128':
            return roofline_0x0_tree_ll128(sizes, n_nodes, nch)
    raise ValueError(f"Unknown config: {split_mask}/{algo}/{proto}")


# ============================================================
# Excel writing
# ============================================================
COLUMNS = ['size_KB', 'Actual Time (us)', 'Roofline Time (us)',
           'Actual/Roofline', 'Actual busBW (GB/s)',
           'Roofline busBW (GB/s)', 'Meas/Roof busBW']


def write_sheet(ws, algo, split_mask, node_scales_data):
    """Write all protocol/channel/node data to a worksheet.

    node_scales_data: dict of {n_nodes: {(algo, proto, nch): (sizes, times, algbws, busbws)}}
    """
    # Clear existing content
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    current_row = 1
    protos = ['Simple', 'LL', 'LL128']

    for n_nodes in sorted(node_scales_data.keys()):
        data = node_scales_data[n_nodes]
        for proto in protos:
            # Get all channel counts for this algo/proto
            channels = sorted([nch for (a, p, nch) in data.keys()
                              if a == algo and p == proto])
            if not channels:
                continue

            for nch in channels:
                key = (algo, proto, nch)
                if key not in data:
                    continue
                sizes, times, algbws, busbws = data[key]

                # Compute roofline
                T_roof, busbw_roof, _ = compute_roofline(
                    split_mask, algo, proto, sizes, n_nodes, nch)

                # Section header
                ws.cell(row=current_row, column=1,
                        value=f"Nodes: {n_nodes}, Protocol: {proto}, Channels: {nch}")
                current_row += 1

                # Column headers
                for col_idx, col_name in enumerate(COLUMNS, 1):
                    ws.cell(row=current_row, column=col_idx, value=col_name)
                current_row += 1

                # Data rows
                for i in range(len(sizes)):
                    size_kb = sizes[i] / 1024
                    actual_time = times[i]
                    roof_time = round(T_roof[i], 4)
                    actual_roof_ratio = round(actual_time / T_roof[i], 4) if T_roof[i] > 0 else 0
                    actual_busbw = busbws[i]
                    roof_busbw = round(busbw_roof[i], 4)
                    meas_roof_bw = round(busbws[i] / busbw_roof[i], 4) if busbw_roof[i] > 0 else 0

                    ws.cell(row=current_row, column=1, value=round(size_kb, 4))
                    ws.cell(row=current_row, column=2, value=actual_time)
                    ws.cell(row=current_row, column=3, value=roof_time)
                    ws.cell(row=current_row, column=4, value=actual_roof_ratio)
                    ws.cell(row=current_row, column=5, value=actual_busbw)
                    ws.cell(row=current_row, column=6, value=roof_busbw)
                    ws.cell(row=current_row, column=7, value=meas_roof_bw)
                    current_row += 1

                # Blank row separator
                current_row += 1

    print(f"  Written {current_row - 1} rows to sheet '{ws.title}'")


# ============================================================
# Main
# ============================================================
def main():
    BASE_DIR = '/Users/liuyaod/summer_2026_lyd/roofline_model'
    EXCEL_PATH = f'{BASE_DIR}/Results.xlsx'

    # Parse all data files
    print("Parsing data files...")
    data_0x7 = {}
    data_0x0 = {}

    for n_nodes in [4, 8, 16]:
        fpath_0x7 = f'{BASE_DIR}/p5en-allreduce-0x7/nccl_test_{n_nodes}nodes_0x7_p5en.out'
        fpath_0x0 = f'{BASE_DIR}/p5en-allreduce-0x0/nccl_test_{n_nodes}nodes_0x0_p5en.out'

        try:
            data_0x7[n_nodes] = parse_file(fpath_0x7)
            print(f"  0x7 {n_nodes} nodes: {len(data_0x7[n_nodes])} sections")
        except FileNotFoundError:
            print(f"  SKIP 0x7 {n_nodes} nodes: file not found")

        try:
            data_0x0[n_nodes] = parse_file(fpath_0x0)
            print(f"  0x0 {n_nodes} nodes: {len(data_0x0[n_nodes])} sections")
        except FileNotFoundError:
            print(f"  SKIP 0x0 {n_nodes} nodes: file not found")

    # Load workbook
    print(f"\nLoading workbook: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH)
    print(f"  Existing sheets: {wb.sheetnames}")

    # Define all sheets to create/update
    sheet_configs = []

    # 0x7 sheets
    for n_nodes in [4, 8, 16]:
        for algo in ['Ring', 'Tree']:
            sheet_name = f'AR, {algo}, 0x7, P5en, {n_nodes}nodes'
            sheet_configs.append({
                'name': sheet_name,
                'algo': algo,
                'split_mask': '0x7',
                'node_scales': [n_nodes],
                'data_source': data_0x7,
            })

    # 0x0 sheets
    for n_nodes in [4, 8, 16]:
        for algo in ['Ring', 'Tree']:
            sheet_name = f'AR, {algo}, 0x0, P5en, {n_nodes}nodes'
            sheet_configs.append({
                'name': sheet_name,
                'algo': algo,
                'split_mask': '0x0',
                'node_scales': [n_nodes],
                'data_source': data_0x0,
            })

    # Process each sheet
    print("\nWriting sheets...")
    for cfg in sheet_configs:
        sheet_name = cfg['name']
        algo = cfg['algo']
        split_mask = cfg['split_mask']
        data_source = cfg['data_source']

        # Collect data for this sheet
        node_scales_data = {}
        for n_nodes in cfg['node_scales']:
            if n_nodes in data_source:
                node_scales_data[n_nodes] = data_source[n_nodes]

        if not node_scales_data:
            print(f"  SKIP {sheet_name}: no data")
            continue

        # Create or get sheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            # Insert new sheet before P6-B200 sheets
            # Find the position of the first P6-B200 sheet
            p6_idx = None
            for i, sn in enumerate(wb.sheetnames):
                if 'P6-B200' in sn:
                    p6_idx = i
                    break
            if p6_idx is not None:
                ws = wb.create_sheet(sheet_name, p6_idx)
            else:
                ws = wb.create_sheet(sheet_name)

        write_sheet(ws, algo, split_mask, node_scales_data)

    # Save
    print(f"\nSaving workbook to: {EXCEL_PATH}")
    wb.save(EXCEL_PATH)
    print("Done!")

    # Print summary
    wb2 = load_workbook(EXCEL_PATH)
    print(f"\nFinal sheets: {wb2.sheetnames}")
    for sn in wb2.sheetnames:
        ws = wb2[sn]
        print(f"  {sn}: {ws.max_row} rows x {ws.max_column} cols")


if __name__ == '__main__':
    main()
